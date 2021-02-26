from tkinter import *
from openpyxl import*

rb = load_workbook("Selection02.xlsx")
sheet = rb.active

numberOfQ: int = sheet.max_row - 1

view = Tk()
view.title("Quiz Application")
view.geometry('1352x652+50+50')
view.configure(bg='bisque')
scrollbar = Scrollbar(view)
scrollbar.pack(side = RIGHT, fill = Y )
T = Text(view, height=1000, width=600,yscrollcommand = scrollbar.set)
T.pack()
numberOfQ: int = sheet.max_row - 1
for row in range(numberOfQ):
    cell = sheet.cell(row=row + 2, column=2)
    T.insert(END, 'Question ' + str(row + 1) + ' ' + cell.value)
    T.insert(END,'\n')
    for i in range(2, 6):
        cell = sheet.cell(row = row + 2, column = i + 1)
        T.insert(END, cell.value)
        T.insert(END, '\n')
    cell = sheet.cell(row=row + 2, column=8)
    T.insert(END,'Your Answer - ' + cell.value)
    T.insert(END, '\n')
    cell = sheet.cell(row=row + 2, column=7)
    T.insert(END,'Correct Answer - ' + cell.value)
    T.insert(END, '\n')
scrollbar.config(command = T.yview )
view.mainloop()