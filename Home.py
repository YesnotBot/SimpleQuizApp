from tkinter import *
from openpyxl import*

rb = load_workbook("Question set.xlsx")
savedfile = "Selection02.xlsx"
sheet = rb.active

numberOfQ: int = sheet.max_row - 1

view = Tk()
view.title("Quiz Application")
view.geometry('1352x652+50+50')
view.configure(bg='bisque')
sideWindow = Frame(view)
sideWindow.pack(side=LEFT)
Qwindow = Frame(view)
Qwindow.pack()
count = 0


def export(val, rowza):
    sheet.cell(row = rowza + 2, column=8).value = val.get()
    rb.save(savedfile)

def submit(view1, vl):
    label = Label(view1, text =vl.get()+" Selected")
    print(vl.get())
    label.pack()
    view1.after(1000, lambda *args: view1.pack_forget())

def submitmain(view1, vl, rowz):
    submit(view1, vl)
    export(vl, rowz)

def nextQn(screen, rowy):
    if rowy < numberOfQ - 1:
        selectQ(screen, rowy+1)
    else :
        label = Label(screen, text = "Question limit over")
        label.pack()

def getview(screen, qS, ans, rowy):
    global count,Qwindow,view,X
    Qwindow.pack_forget()
    count=count+1
    Qwindow = Frame(view)
    v = StringVar(Qwindow, "1")
    label = Label(Qwindow, text = qS)
    button_a = Radiobutton(Qwindow, text = ans[0], variable = v, value = "A", activebackground = "goldenrod", bg="seagreen3", indicator = 0)
    button_b = Radiobutton(Qwindow, text = ans[1], variable = v, value = "B", activebackground = "goldenrod", bg="seagreen3", indicator = 0)
    button_c = Radiobutton(Qwindow, text = ans[2], variable = v, value = "C", activebackground = "goldenrod", bg="seagreen3", indicator = 0)
    button_d = Radiobutton(Qwindow, text = ans[3], variable = v, value = "D", activebackground = "goldenrod", bg="seagreen3", indicator = 0)
    label.pack()
    button_a.pack(side = "top")
    button_b.pack(side = "top")
    button_c.pack(side = "top")
    button_d.pack(side = "top")
    button_submit = Button(Qwindow, text = "Submit", activebackground = "coral", bg="rosy brown", command = lambda *args: submitmain(Qwindow, v, rowy))
    button_submit.pack(side = "left")
    button_next = Button(Qwindow, text ="Next", activebackground = "coral", bg="indian red", command = lambda *args: nextQn(Qwindow, rowy))
    button_next.pack(side = "right")
    Qwindow.pack()




def selectQ(screen, rowx):
    global X
    answer = []
    questionString = " "
    X.configure(bd=0)
    cell = sheet.cell(row=rowx + 2, column=2)
    questionString = cell.value
    for i in range(2, 6):
        cell = sheet.cell(row = rowx + 2, column = i + 1)
        answer.append(cell.value)
    getview(screen, questionString, answer, rowx)
    answer = []

def combinecall(Qwindow, rowy):
    Qwindow.pack_forget()
    Qwindow.pack()
    selectQ(Qwindow, rowy)


"""" def refreshWindow(SW, viewy, nQ):
    row = 0
    command = (lambda viewy, row: lambda: selectQ(viewy, row))(viewy, row)
    Q1 = Button(SW, text="Q1", bg="Skyblue1", bd=0, height=3, width=4, command=command)
    Q1.grid()
    """

row = 0


for row in range(numberOfQ):
    command = (lambda viewy, row: lambda: combinecall(viewy, row))(Qwindow, row)
    X = Button(sideWindow, text="Q" + str(row + 1), bg="Skyblue1", height=3, width=4, command=command)
    X.grid()
# refreshWindow(sideWindow, view, numberOfQ)

rb.save(savedfile)
view.mainloop()
